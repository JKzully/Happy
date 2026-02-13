"""
Import "Sölutölur Innslegnar.xlsx" directly into Supabase daily_sales table.
Reads the Excel, maps product names → product UUIDs, store names → store UUIDs,
and upserts all sales rows.
"""
import os
import sys
import json
import requests
import openpyxl

SUPABASE_URL = "https://cjdcxzdjdmycanhkgphp.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNqZGN4emRqZG15Y2FuaGtncGhwIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MDU0NjcxMywiZXhwIjoyMDg2MTIyNzEzfQ._rYiRxQYc79wjseGmiGEWPEERQJxUb_kqC-dvFSGLNQ"

HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates",
}

# Product name → product slug mapping (matches sku-map.ts productNameToId)
PRODUCT_NAME_MAP = {
    "lemon lime": "lemon-lime",
    "mixed berries": "mixed-berries",
    "pina colada": "pina-colada",
    "peach": "peach",
    "peru": "peru",
    "creatine lemon": "creatine-lemon",
    "creatine mixed": "creatine-mixed",
    "energy apple kiwi": "energy-kiwi",
    "kids apple kiwi": "krakka-green-apple-kiwi",
    "kids mixed berries": "krakka-mixed-berry",
    "jóli": "jolabragd",
    "joli": "jolabragd",
}

# Chain sheet → chain slug
CHAIN_MAP = {
    "Krónan": "kronan",
    "Krónan ": "kronan",
    "Bónus": "bonus",
    "Hagkaup": "hagkaup",
    "Samkaup": "samkaup",
}

def supabase_get(table, params=""):
    url = f"{SUPABASE_URL}/rest/v1/{table}?{params}"
    r = requests.get(url, headers=HEADERS)
    r.raise_for_status()
    return r.json()

def supabase_post(table, rows):
    # on_conflict ensures upsert on the unique constraint instead of 409 errors
    url = f"{SUPABASE_URL}/rest/v1/{table}?on_conflict=date,store_id,product_id,order_type"
    r = requests.post(url, headers=HEADERS, json=rows)
    if r.status_code not in (200, 201):
        print(f"  ERROR: {r.status_code} {r.text[:200]}")
    return r

def main():
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "C:/Users/solvi/hce-dashboard/innslegnar.xlsx"
    print(f"Reading {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # 1. Fetch products from DB
    print("Fetching products from Supabase...")
    products = supabase_get("products", "select=id,name")
    product_name_to_uuid = {}
    for p in products:
        product_name_to_uuid[p["name"].lower()] = p["id"]
    print(f"  Found {len(products)} products: {[p['name'] for p in products]}")

    # Build slug → UUID map using the same slugs as PRODUCT_NAME_MAP values
    product_slug_to_uuid = {
        "lemon-lime": None, "mixed-berries": None, "pina-colada": None,
        "peach": None, "peru": None, "creatine-lemon": None,
        "creatine-mixed": None, "energy-kiwi": None,
        "krakka-green-apple-kiwi": None, "krakka-mixed-berry": None,
        "jolabragd": None,
    }
    # Match by slugifying the DB product name
    for p in products:
        slug = p["name"].lower().replace(" ", "-").replace("ó", "o").replace("ð", "d")
        for key in product_slug_to_uuid:
            normalized_key = key.replace("ó", "o").replace("ð", "d")
            if slug == normalized_key or slug == key:
                product_slug_to_uuid[key] = p["id"]
                break
        # Also try direct lowercase match
        product_slug_to_uuid[p["name"].lower().replace(" ", "-")] = p["id"]
    # Manual overrides for tricky names
    for p in products:
        if "ólabrag" in p["name"].lower():
            product_slug_to_uuid["jolabragd"] = p["id"]
        if "krakka" in p["name"].lower() or "kids green" in p["name"].lower():
            product_slug_to_uuid["krakka-green-apple-kiwi"] = p["id"]
        if "kids mixed" in p["name"].lower():
            product_slug_to_uuid["krakka-mixed-berry"] = p["id"]
    print(f"  Slug map: {product_slug_to_uuid}")

    # 2. Fetch chains from DB
    print("Fetching chains from Supabase...")
    chains = supabase_get("retail_chains", "select=id,slug")
    chain_slug_to_uuid = {c["slug"]: c["id"] for c in chains}
    print(f"  Found chains: {chain_slug_to_uuid}")

    # 3. Fetch stores from DB
    print("Fetching stores from Supabase...")
    stores = supabase_get("stores", "select=id,name,chain_id,sub_chain_type")
    # Build lookup: (chain_uuid, store_name_lower) → store_uuid
    store_lookup = {}
    for s in stores:
        store_lookup[(s["chain_id"], s["name"].lower().strip())] = s["id"]
    print(f"  Found {len(stores)} stores")

    # 4. Process each chain sheet
    total_inserted = 0
    total_skipped = 0
    missing_products = set()
    missing_stores = set()

    for sheet_name, chain_slug in CHAIN_MAP.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        chain_uuid = chain_slug_to_uuid.get(chain_slug)
        if not chain_uuid:
            print(f"  SKIP: Chain '{chain_slug}' not in DB")
            continue

        print(f"\nProcessing {sheet_name.strip()} ({ws.max_row} rows)...")

        # Read header row to get store names
        header = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(1, c).value
            header.append(str(v).strip() if v else "")

        # Store columns start at index 2 (col C)
        store_cols = []  # (col_index, store_name, store_uuid)
        for i in range(2, len(header)):
            store_name = header[i].strip()
            if not store_name:
                continue

            # For Samkaup, store names are like "Nettó-Borgarnes", "Kjörbúð-Blönduós"
            # Try exact match first, then try variations
            store_uuid = None
            name_lower = store_name.lower().strip()

            # Try exact match
            store_uuid = store_lookup.get((chain_uuid, name_lower))

            # Try with chain prefix removed for Samkaup
            if not store_uuid and chain_slug == "samkaup":
                # "Nettó-Borgarnes" → try "Nettó Borgarnes" and "Borgarnes"
                parts = store_name.split("-", 1)
                if len(parts) == 2:
                    sub_chain = parts[0].strip()
                    location = parts[1].strip()
                    # Try "SubChain Location"
                    store_uuid = store_lookup.get((chain_uuid, f"{sub_chain} {location}".lower()))
                    # Try just location
                    if not store_uuid:
                        store_uuid = store_lookup.get((chain_uuid, location.lower()))
                    # Try with "Kjörbúðin" prefix
                    if not store_uuid and sub_chain.lower().startswith("kjörbúð"):
                        store_uuid = store_lookup.get((chain_uuid, f"kjörbúðin {location}".lower()))

            # Try partial match
            if not store_uuid:
                for (cid, sname), sid in store_lookup.items():
                    if cid == chain_uuid:
                        if sname.startswith(name_lower[:4]) or name_lower.startswith(sname[:4]):
                            store_uuid = sid
                            break

            if store_uuid:
                store_cols.append((i, store_name, store_uuid))
            else:
                missing_stores.add(f"{sheet_name.strip()}: {store_name}")

        print(f"  Matched {len(store_cols)} stores, {len([s for s in header[2:] if s.strip()]) - len(store_cols)} unmatched")

        # Read data rows (collect and deduplicate)
        dedup = {}  # (date, store_id, product_id) → row dict

        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            product_val = ws.cell(r, 2).value

            if not date_val or not product_val:
                continue

            product_name = str(product_val).strip()
            if product_name.lower() == "samtals":
                continue

            # Parse date
            date_str = str(date_val)[:10]
            if "00:00:00" in str(date_val):
                date_str = str(date_val)[:10]

            # Map product name
            product_slug = PRODUCT_NAME_MAP.get(product_name.lower())
            if not product_slug:
                missing_products.add(product_name)
                continue

            # Look up UUID from slug map
            product_uuid = product_slug_to_uuid.get(product_slug)
            if not product_uuid:
                missing_products.add(f"{product_name} (slug: {product_slug})")
                continue

            # Read quantities for each store
            for col_idx, store_name, store_uuid in store_cols:
                qty_val = ws.cell(r, col_idx + 1).value  # openpyxl is 1-indexed
                if qty_val is None or qty_val == "" or qty_val == 0:
                    continue
                try:
                    qty = int(float(str(qty_val)))
                except (ValueError, TypeError):
                    continue
                if qty <= 0:
                    continue

                key = (date_str, store_uuid, product_uuid)
                # Deduplicate: sum quantities for same date+store+product
                if key in dedup:
                    dedup[key]["quantity"] += qty
                else:
                    dedup[key] = {
                        "date": date_str,
                        "store_id": store_uuid,
                        "product_id": product_uuid,
                        "quantity": qty,
                        "order_type": "retail",
                    }

        # Insert deduped rows in batches of 500
        all_rows = list(dedup.values())
        print(f"  {len(all_rows)} unique rows after dedup")
        for i in range(0, len(all_rows), 500):
            batch = all_rows[i:i+500]
            supabase_post("daily_sales", batch)
            total_inserted += len(batch)
            print(f"  Inserted {total_inserted} rows...", end="\r")

        print(f"  Inserted {len(all_rows)} rows for {sheet_name.strip()}")

    # Summary
    print(f"\n{'='*50}")
    print(f"DONE: {total_inserted} rows inserted")

    if missing_products:
        print(f"\nMissing products ({len(missing_products)}):")
        for p in sorted(missing_products):
            print(f"  - {p}")

    if missing_stores:
        print(f"\nUnmatched stores ({len(missing_stores)}):")
        for s in sorted(missing_stores):
            print(f"  - {s}")

if __name__ == "__main__":
    main()
